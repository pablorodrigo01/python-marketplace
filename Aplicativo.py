import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
import pandas as pd
import json
import datetime
from openpyxl import load_workbook

# Função para ler o arquivo de configuração
def read_config(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        config = json.load(file)
    return config

# Função para criar o arquivo log.txt
def generate_log(platform, sku):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_message = f"[{timestamp}] SKU não encontrado na tabela de entrada - Plataforma: {platform}, SKU: {sku}\n"

    with open('log.txt', 'a+', encoding='utf-8') as log_file:
        log_file.seek(0)
        logs = log_file.read()

        if log_message not in logs:
            log_file.write(log_message)

# Carrega as configurações do arquivo
config = read_config('config.json')
platform_configs = config['platforms']

# Função para fechar a janela corretamente
def close():
    app.destroy()

# Função para atualizar o arquivo em todas as plataformas
def update_all():
    file_path = file_input.get()

    try:
        # Lê o arquivo de entrada e extrai as colunas SKU, Nome, Preço e Estoque
        df = pd.read_excel(file_path, usecols='A, B, C, D', skiprows=1)
        df['Nome'] = df['Nome'].str.upper()

        # Lista para armazenar os SKUs da tabela de entrada
        input_skus = df['Codigo'].tolist()

        for config in platform_configs:
            platform = config['name']
            sheet_name = config['sheet']
            e_col = config['e']
            sku_col = config['sku']
            p_col = config['p']
            file_output = config['file_output']

            # Abre o arquivo de destino em modo de leitura e escrita
            book = load_workbook(file_output)
            sheet = book[sheet_name]

            # Lista para armazenar os SKUs da plataforma
            platform_skus = []

            for i in range(4, sheet.max_row + 1):
                sku_dest = sheet.cell(row=i, column=sku_col).value
                platform_skus.append(sku_dest)

            # Comparar os SKUs da plataforma com os da tabela de entrada
            skus_not_in_input = [sku for sku in platform_skus if sku is not None and sku not in input_skus]

            # Imprimir os SKUs que não estão na tabela de entrada
            for sku in skus_not_in_input:
                generate_log(platform, sku)

            if platform == 'Mercado Livre' or platform == 'Shopee':
                for _, row in df.iterrows():
                    sku = row['Codigo']
                    preco = row['Custo']
                    estoque = row['Estoque']

                    for i in range(4, sheet.max_row + 1):
                        sku_dest = sheet.cell(row=i, column=sku_col).value
                        if str(sku_dest) == str(sku):
                            sheet.cell(row=i, column=e_col).value = estoque
                            sheet.cell(row=i, column=p_col).value = preco
                            if platform == 'Mercado Livre':
                                sheet.cell(row=i, column=9).value = preco

            if platform == 'Amazon':
                # Abre o arquivo de destino em modo de leitura e escrita
                book = load_workbook(file_output)
                sheet = book[sheet_name]

                # Limpar a planilha da Amazon, mantendo apenas o cabeçalho
                for i in range(sheet.max_row, 1, -1):
                    if i > 1:
                        sheet.delete_rows(i)

                # Verifica a próxima linha disponível (depois da limpeza)
                next_row = 2  # Começando da segunda linha após o cabeçalho

                # Iterar sobre as linhas da tabela do Mercado Livre
                for _, row_ml in df.iterrows():
                    sku_ml = row_ml['Codigo']
                    preco_ml = row_ml['Custo']
                    estoque_ml = row_ml['Estoque']

                    # Insira os dados do Mercado Livre diretamente nas colunas apropriadas da tabela da Amazon
                    sheet.cell(row=next_row, column=sku_col).value = sku_ml
                    sheet.cell(row=next_row, column=p_col).value = preco_ml
                    sheet.cell(row=next_row, column=e_col).value = estoque_ml

                    # Atualize a próxima linha disponível
                    next_row += 1

                # Salve as alterações no arquivo
                book.save(file_output)

        # Mostra uma mensagem de sucesso
        messagebox.showinfo("Sucesso", "Arquivo atualizado com sucesso!")

    except Exception as e:
        # Mostra uma mensagem de erro
        messagebox.showerror("Erro", str(e))
# Cria uma janela
app = ctk.CTk()
app.title("Reman Aplicativo - Marketplace")

ctk.set_appearance_mode("dark")  # Modes: system (default), light, dark
ctk.set_default_color_theme("blue")  # Themes: blue (default), dark-blue, green

# Define o tamanho e posição da janela
windows_width = 800
windows_height = 250
size_width = app.winfo_screenwidth()
size_height = app.winfo_screenheight()
posx = size_width/2 - windows_width/2
posy = size_height/2 - windows_height/2
app.geometry("%dx%d+%d+%d" % (windows_width, windows_height, posx, posy))

# Define a função de fechamento da janela quando clicar no botão "X"
app.protocol("WM_DELETE_WINDOW", close)

# Frame para o cabeçalho
header_frame = ctk.CTkFrame(app)
header_frame.pack(pady=20)

# Label do cabeçalho
header_label = ctk.CTkLabel(master=header_frame, text="Atualização de Tabelas", font=("Myriad Pro", 20))
header_label.pack()

# Frame para o conteúdo
content_frame = ctk.CTkFrame(app)
content_frame.pack(pady=20)

# Cria os widgets para seleção dos arquivos
file_label = ctk.CTkLabel(content_frame, text="Selecione o arquivo de entrada:", font=("Myriad Pro", 16))
file_label.grid(row=2, column=0, padx=20, pady=10)

file_input_frame = ctk.CTkFrame(content_frame)
file_input_frame.grid(row=2, column=1, padx=10, pady=10)

file_input = ctk.CTkEntry(file_input_frame, width=200, font=("Myriad Pro", 16))
file_input.grid(row=0, column=0)

browse_input_button = ctk.CTkButton(file_input_frame, text="Procurar", font=("Myriad Pro", 16), command=lambda: file_input.insert(tk.END, filedialog.askopenfilename()))
browse_input_button.grid(row=0, column=1, padx=10)

# Botão para atualizar o arquivo em todas as plataformas
update_button = ctk.CTkButton(content_frame, text="Atualizar Arquivo", font=("Myriad Pro", 16), command=update_all)
update_button.grid(row=4, column=0, columnspan=2, pady=10)

# Inicia o loop principal da interface gráfica
app.mainloop()