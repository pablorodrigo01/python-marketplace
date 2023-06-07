import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
import pandas as pd
from openpyxl import load_workbook

# Função para fechar a janela corretamente
def close():
    app.destroy()

# Função para atualizar o arquivo em todas as plataformas
def update_all():
    file_path = file_input.get()

    try:
        # Lê o arquivo de entrada e extrai as colunas SKU, Nome, Preço e Estoque
        df = pd.read_excel(file_path, usecols='A, B, C, D', skiprows=1)
        # Instrução passada para definir todos os itens da coluna "Nome" em uppercase
        df['Nome'] = df['Nome'].str.upper()

        # Configurações das plataformas
        # Legenda: cn -> Coluna Nome, ce -> Coluna Estoque, cp -> Coluna Preço, file_output -> Nome do arquivo de saída
        platform_configs = [
            {
                'name': 'Mercado Livre',
                'sheet': 'Anúncios',
                'cn': 4,
                'sku': 3,
                'ce': 6,
                'cp': 8,
                'file_output': 'file\\MercadoLivre.xlsx'
            },
            {
                'name': 'Shopee',
                'sheet': 'Sheet1',
                'cn': 2,
                'sku': 5,
                'ce': 9,
                'cp': 7,
                'file_output': 'file\\Shopee.xlsx'
            },
            {
                'name': 'Magalu',
                'sheet': 'Cadastro de Produtos SEM VARIAÇ',
                'cn': 2,
                'ce': 4,
                'cp': 3,
                'sku': None,
                'file_output': 'file\\Magalu.xlsx'
            }
        ]

        for config in platform_configs:
            platform = config['name']
            sheet_name = config['sheet']
            cn_col = config['cn']
            ce_col = config['ce']
            sku_col = config['sku']
            cp_col = config['cp']
            file_output = config['file_output']

            # Abre o arquivo de destino em modo de leitura
            book = load_workbook(file_output)

            # Seleciona a planilha para atualização
            sheet = book[sheet_name]

            # Itera sobre as linhas do arquivo de entrada
            for _, row in df.iterrows():
                sku = row['SKU']
                nome = row['Nome']
                preco = row['Custo']
                estoque = row['Estoque']

                if sku_col is not None:
                    # Procura o SKU correspondente na planilha de destino
                    for i in range(4, sheet.max_row + 1):
                        sku_dest = sheet.cell(row=i, column=sku_col).value
                        if str(sku_dest) == str(sku):
                            # Atualiza a coluna de Nome na planilha de destino
                            sheet.cell(row=i, column=cn_col).value = nome
                            break

                # Procura a linha correspondente na planilha de destino
                for j in range(4, sheet.max_row + 1):
                    titulo = sheet.cell(row=j, column=cn_col).value
                    if titulo == nome:
                        # Atualiza as células correspondentes na planilha de destino
                        sheet.cell(row=j, column=cp_col).value = estoque
                        sheet.cell(row=j, column=cp_col).value = preco
                        
                        # Atualiza somente a coluna 9 da planilha do Mercado Livre com o valor de preco
                        if platform == 'Mercado Livre':
                            sheet.cell(row=j, column=9).value = preco
                        
                        break

            # Salva as alterações no arquivo de destino
            book.save(file_output)

        # Mostra uma mensagem de sucesso
        messagebox.showinfo("Sucesso", "Arquivo atualizado com sucesso!")

    except Exception as e:
        # Mostra uma mensagem de erro
        messagebox.showerror("Erro", str(e))

# Cria uma janela
app = ctk.CTk()
app.title("Reman Aplicativo - Marketplace")

ctk.set_appearance_mode("dark")  # Modes: system (default), light, dark
ctk.set_default_color_theme("blue")  # Themes: blue (default), dark-blue, green

# Define o tamanho e posição da janela
largura_janela = 800
altura_janela = 250
largura_tela = app.winfo_screenwidth()
altura_tela = app.winfo_screenheight()
posx = largura_tela/2 - largura_janela/2
posy = altura_tela/2 - altura_janela/2
app.geometry("%dx%d+%d+%d" % (largura_janela, altura_janela, posx, posy))

# Define a função de fechamento da janela quando clicar no botão "X"
app.protocol("WM_DELETE_WINDOW", close)

# Frame para o cabeçalho
header_frame = ctk.CTkFrame(app)
header_frame.pack(pady=20)

# Label do cabeçalho
header_label = ctk.CTkLabel(master=header_frame, text="Atualização de Tabelas", font=("Myriad Pro", 20))
header_label.pack()

# Frame para o conteúdo
content_frame = ctk.CTkFrame(app)
content_frame.pack(pady=20)

# Cria os widgets para seleção dos arquivos
file_label = ctk.CTkLabel(content_frame, text="Selecione o arquivo de entrada:", font=("Myriad Pro", 16))
file_label.grid(row=2, column=0, padx=20, pady=10)

file_input_frame = ctk.CTkFrame(content_frame)
file_input_frame.grid(row=2, column=1, padx=10, pady=10)

file_input = ctk.CTkEntry(file_input_frame, width=200, font=("Myriad Pro", 16))
file_input.grid(row=0, column=0)

browse_input_button = ctk.CTkButton(file_input_frame, text="Procurar", font=("Myriad Pro", 16), command=lambda: file_input.insert(
    tk.END, filedialog.askopenfilename()))
browse_input_button.grid(row=0, column=1, padx=10)

# Botão para atualizar o arquivo em todas as plataformas
atualizar_button = ctk.CTkButton(content_frame, text="Atualizar Arquivo", font=("Myriad Pro", 16), command=update_all)
atualizar_button.grid(row=4, column=0, columnspan=2, pady=10)

# Inicia o loop principal da interface gráfica
app.mainloop()
